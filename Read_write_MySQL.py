import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)
from openpyxl.utils.dataframe import dataframe_to_rows

# Some sample data to plot.
list_data = [111, 20, 30, 20, 15, 30, 62]

# Create a Pandas dataframe from the data.
df = pd.DataFrame(list_data)

# Define the file name and sheet name.
excel_file = '/Users/michaelaneiro/Downloads/column.xlsx'
sheet_name = 'Sheet1'

# Try to load the workbook. If it doesn't exist, create a new workbook.
try:
    print("Load the workbook using openpyxl")
    wb = load_workbook(filename=excel_file)
except FileNotFoundError:
    print("If the file does not exist, create a new workbook")
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name

# Create the worksheet object.
ws = wb[sheet_name]

# Write the new data to the worksheet.
for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# Save the workbook.
wb.save(excel_file)