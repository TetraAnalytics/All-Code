from openpyxl import load_workbook, Workbook
from os.path import exists


def print_mortgage_results():
    print("PRINT DIRECTLY TO CELLS IN AN XLSX SHEET")
    # Define your variables
    variable1 = "Loan id"
    variable2 = "Price"
    variable3 = "Accrued"
    variable4 = "Value"
    variable5 = "Modified Duration"
    variable6 = "Weighted Avg Life"
    variable7 = "Yield"

    # Define the cells you want to write the variables to
    cell1 = "A1"
    cell2 = "B1"
    cell3 = "C1"
    cell4 = "D1"
    cell5 = "E1"
    cell6 = "F1"
    cell7 = "G1"

    # Define the path to your Excel file and the sheet name
    mortgage_results = '/Users/michaelaneiro/Downloads/mortgage_results_out.xlsx'
    sheet_name = 'Results'

    # Check if the file exists
    if not exists(mortgage_results):
        # If the file does not exist, create a new workbook and add a sheet with the specified name
        print("file does not exist")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
    else:
        # If the file exists, load the workbook and select the sheet
        print("file exists")
        workbook = load_workbook(mortgage_results)
        sheet_name = "Results"
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            sheet = workbook.create_sheet(sheet_name)

    # Write the variables to the specified cells
    sheet[cell1] = variable1  # m.loan_id
    sheet[cell2] = variable2  # round(m.price - m.accrued_interest, 3)
    sheet[cell3] = variable3  # round(m.accrued_interest, 3)
    sheet[cell4] = variable4  # round(m.price, 3)
    sheet[cell5] = variable5  # round(m.modified_duration, 2)
    sheet[cell6] = variable6  # round(m.weighted_average_life, 2)
    sheet[cell7] = variable7  # m.BEY

    # Write the variables to the specified cells
    sheet["A3"] = "m.loan_id"
    sheet["B3"] = "round(m.price - m.accrued_interest, 3)"
    sheet["C3"] = "DONT REPLACE"
    sheet["D3"] = "round(m.price, 3)"
    sheet["E3"] = "round(m.modified_duration, 2)"
    sheet["F3"] = "round(m.weighted_average_life, 2)"
    sheet["G3"] = "m.BEY"

    print("Sheet Name ", sheet_name)
    for row in sheet.iter_rows(min_row=1, max_col=3, max_row=3):
        for cell in row:
            print(cell.value)

    # looping through each row and column
    for x in range(1, 5):
        for y in range(1, 5):
            print(x, y, sheet.cell(row=x, column=y).value)

    # iterating rows
    for row in sheet.iter_rows(min_row=2, max_col=3, max_row=3):
        for cell in row:
            print(cell.value)

    # Save the changes to the workbook
    workbook.save(mortgage_results)
    print("Pricing Results")

    return


print_mortgage_results()
