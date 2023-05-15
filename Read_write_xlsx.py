from openpyxl import Workbook, load_workbook
import os
from os.path import exists
import pandas as pd


def print_mortgage_results(m):
    print("PRINT DIRECTLY TO CELLS IN AN XLSX SHEET")
    # Define your variables
    variable1 = "Loan id"
    variable2 = "Price"
    variable3 = "Accrued"
    variable4 = "Value"
    variable5 = "Modified Duration"
    variable6 = "Weighted Avg Life"
    variable7 = "Yield"

    # Define the cells you want to write the variables to
    cell1 = "A1"
    cell2 = "B1"
    cell3 = "C1"
    cell4 = "D1"
    cell5 = "E1"
    cell6 = "F1"
    cell7 = "G1"

    # Define the path to your Excel file and the sheet name
    mortgage_results = '/Users/michaelaneiro/Downloads/mortgage_results_out.xlsx'
    sheet_name = 'Results'

    # Check if the file exists
    if not exists(mortgage_results):
        # If the file does not exist, create a new workbook and add a sheet with the specified name
        print("file does not exist")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
    else:
        # If the file exists, load the workbook and select the sheet
        print("file exists")
        workbook = load_workbook(mortgage_results)
        sheet_name = "Results"
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            sheet = workbook.create_sheet(sheet_name)

    # Write the variables to the specified cells
    sheet[cell1] = variable1
    sheet[cell2] = variable2
    sheet[cell3] = variable3
    sheet[cell4] = variable4
    sheet[cell5] = variable5
    sheet[cell6] = variable6
    sheet[cell7] = variable7

    # Write the variables to the specified cells
    sheet["A2"] = m.loan_id
    sheet["B2"] = round(m.price - m.accrued_interest, 3)
    sheet["C2"] = round(m.accrued_interest, 3)
    sheet["D2"] = round(m.price, 3)
    sheet["E2"] = round(m.modified_duration, 2)
    sheet["F2"] = round(m.weighted_average_life, 2)
    sheet["G2"] = m.BEY * 100.0

    # Save the changes to the workbook
    workbook.save(mortgage_results)
    print("Pricing Results")

    return


def print_mortgage_cashflows_xlsx(m):  # pass the Mortgage object and the iteration of loan analysis
    # Open a new XLSX file in write mode
    mortgage_results = '/Users/michaelaneiro/Downloads/MicroHubCashFlows.xlsx'

    # Create a new DataFrame to store the results
    df = pd.DataFrame(columns=['Balance', 'Scheduled Pay', 'Scheduled Prin', 'Scheduled Interest', 'Prepay', 'Default', 'Loss'])
    for i in range(m.amortization_term):  # After the header defines the df fields, load df data from each month's cashflow
        df.loc[i] = [m.cashflows[i].balance, m.cashflows[i].scheduled_payment, m.cashflows[i].scheduled_principal,  m.cashflows[i].scheduled_interest, m.cashflows[i].prepayment_principal, m.cashflows[i].default_principal, m.cashflows[i].default_loss]

    # if exists(mortgage_results):  # temporary cheat until I correct lower code logic
      #  os.remove(mortgage_results)

    # Check if the file exists
    if not exists(mortgage_results):
        # If the file doesn't exist, create a new Excel file with the specific sheet name
        print("Mortgage_Cashflows New")
        sheet_name = "HomeLoans"
        with pd.ExcelWriter(mortgage_results, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # If the file exists, update the existing Excel file with the new data
        print("Mortgage_Cashflows Exists")
        sheet_name = "HomeLoans"
        book = None
        workbook = load_workbook(mortgage_results)
        try:
            sheet = workbook[sheet_name]
            with pd.ExcelWriter(mortgage_results, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except KeyError:
            sheet = workbook.create_sheet(sheet_name)
            with pd.ExcelWriter(mortgage_results, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    return
